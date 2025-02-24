from openpyxl import Workbook
from openpyxl.styles import Font

grid = [
    # 1. Seite
    ["s", "ui", "e"],
    ["e", "s", "ui"],
    ["ui", "e", "s"],
    ["s", "ui", "e"],
    ["e", "s", "ui"],
    ["ui", "e", "s"],
    ["s", "ui", "e"],
    # 2. Seite
    ["ui", "e", "s"],
    ["e", "s", "ui"],
    ["s", "ui", "e"],
    ["e", "ui", "s"],
    ["ui", "e", "s"],
    ["s", "e", "ui"],
    ["e", "ui", "s"],
    # 3. Seite
    ["e", "s", "ui"],
    ["s", "e", "ui"],
    ["e", "ui", "s"],
    ["ui", "s", "e"],
    ["ui", "s", "e"],
    ["ui", "e", "s"],
    ["s", "e", "ui"],
    # 4. Seite
    ["ui", "s", "e"],
    ["s", "e", "ui"],
    ["ui", "s", "e"],
    ["e", "ui", "s"],
    ["s", "e", "ui"],
    ["s", "ui", "e"],
    ["e", "ui", "s"],
]

facetten = [
    "Fähigkeit zur Einschätzung des eigenen Lernstands",
    "Fähigkeit adäquate Lernziele zu setzen",
    "Wahl einer geeigneten Lernstrategie",
    "Anwendungsgüte der Lernstrategie",
    "Fähigkeit zur Feststellung des eigenen Lernfortschritts",
    "Fähigkeit zur Anpassung des eigenen Lernens",
    "Überprüfung und Feststellung des Lernergebnisses",
]


def get_facette(facette, row):
    result = {"s": 0, "ui": 0, "e": 0}

    for i in range(facette, len(row), 7):
        antwortABC = row[i].value
        if antwortABC:
            antwortNumber = ord(antwortABC) - ord("A")
            # Antwort basierend auf dem Grid auswerten
            antwort = grid[i - 1][antwortNumber]
            result[antwort] += 1

    return result


def get_final_result(facette):
    if facette["s"] == 4:
        return "Selbstreguliert"
    if facette["s"] == 3:
        return "Überwiegend selbstreguliert"

    if facette["s"] == 2:
        if facette["e"] == 1 and facette["ui"] == 1:
            return "Ansatzweise selbstreguliert"
        if facette["e"] == 2:
            return "Mischtyp selbstreguliert / external reguliert"
        if facette["ui"] == 2:
            return "Mischtyp selbstreguliert / unreflektiert-impulsiv"

    if facette["s"] == 1:
        if facette["e"] == 3:
            return "Überwiegend external reguliert"
        if facette["ui"] == 3:
            return "Überwiegend unreflektiert-impulsiv"
        if facette["e"] == 2 and facette["ui"] == 1:
            return "Ansatzweise external reguliert"
        if facette["e"] == 1 and facette["ui"] == 2:
            return "Ansatzweise unreflektiert-impulsiv"

    if facette["s"] == 0:
        if facette["e"] == 4:
            return "External reguliert"
        if facette["ui"] == 4:
            return "Unreflektiert-impulsiv"
        if facette["e"] == 2 and facette["ui"] == 2:
            return "Mischtyp external reguliert / unreflektiert-impulsiv"
        if facette["e"] == 3 and facette["ui"] == 1:
            return "Überwiegend external reguliert"
        if facette["e"] == 1 and facette["ui"] == 3:
            return "Überwiegend unreflektiert-impulsiv"

    return f"Keine Zuordnung für diesen Fall ({facette})"


def evaluate_results(workbook):
    sheet = workbook.active

    # Neue Excel-Datei erstellen
    result_workbook = Workbook()

    # Iteriere über jede Zeile (ab der zweiten Zeile für die Teilnehmer)
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        teilnehmer_name = row[0].value  # Name des Teilnehmers aus Spalte A
        if not teilnehmer_name:
            continue

        # Neues Sheet für den Teilnehmer erstellen
        result_sheet = result_workbook.create_sheet(title=teilnehmer_name[:30])  # max. 31 Zeichen für Sheet-Namen

        # Überschriften einfügen und formatieren
        result_sheet.append(["Facette", "Details", "Bewertung"])
        # Überschriftenzeile fett machen
        for cell in result_sheet[1]:  # Zeile 1 ist die erste Zeile (1-basiert)
            cell.font = Font(bold=True)

        # Ergebnisse für jede Facette berechnen
        for facette_index in range(7):
            facette_result = get_facette(facette_index + 1, row)
            final_result = get_final_result(facette_result)

            # Ergebnisse in das Teilnehmer-Sheet schreiben
            result_sheet.append([facetten[facette_index], str(facette_result), final_result])

        # Spaltenbreiten anpassen für bessere Lesbarkeit
        result_sheet.column_dimensions['A'].width = 50  # Facetten-Spalte
        result_sheet.column_dimensions['B'].width = 20  # Ergebnisse-Spalte
        result_sheet.column_dimensions['C'].width = 50  # Bewertungsspalte

    # Standardsheet entfernen
    if "Sheet" in result_workbook.sheetnames:
        del result_workbook["Sheet"]

    return result_workbook