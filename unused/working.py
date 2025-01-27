from openpyxl import load_workbook, Workbook


def evaluate_answers(file_path, sheet_name, row_number):
    # Excel-Datei laden
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Teilnehmername aus Spalte H
    participant_name = sheet[f"H{row_number}"].value

    # Startspalte für die Antworten (Frage 2 beginnt bei Spalte T)
    start_column = ord("T")  # Spalte T als Basis
    num_questions = 28
    offset = 7  # Abstand der Fragen
    choices = 3  # Anzahl der Antwortmöglichkeiten (A, B, C)

    answers = []  # Hier speichern wir die Antworten ('A', 'B' oder 'C')

    # Schleife über alle Fragen
    for question in range(1, num_questions + 1):
        # Spalten der Antwortmöglichkeiten (A, B, C)
        answer_columns = []
        for i in range(choices):
            col_index = start_column + (question - 1) * offset + i
            first_letter = chr(ord("A") + (col_index - ord("A")) // 26 - 1)
            second_letter = chr(ord("A") + (col_index - ord("A")) % 26)
            if first_letter < "A":
                answer_columns.append(second_letter)
            else:
                answer_columns.append(first_letter + second_letter)

        # Werte der Zellen abrufen
        values = [sheet[f"{col}{row_number}"].value for col in answer_columns]

        # Prüfen, welche Antwort gewählt wurde (1 steht für die Auswahl)
        if values[0] == 1:
            answers.append("A")
        elif values[1] == 1:
            answers.append("B")
        elif values[2] == 1:
            answers.append("C")
        else:
            answers.append(None)  # Für den Fall, dass keine gültige Antwort vorliegt

    workbook.close()

    # Rückgabe des Teilnehmernamens und der Antworten
    return participant_name, answers


def save_results(output_file, participant_name, answers):
    # Neue Excel-Datei erstellen
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Header hinzufügen
    sheet.append(["Teilnehmer", *[f"Frage {i+1}" for i in range(len(answers))]])

    # Daten hinzufügen
    sheet.append([participant_name, *answers])

    # Datei speichern
    workbook.save(output_file)
    print(f"Ergebnis wurde in '{output_file}' gespeichert.")


# Beispielaufruf
file_path = "data.xlsx"
sheet_name = "Grundtypen_1.Tabelle_1.Versuch"
row_number = 3  # Zeile 3 auswerten
output_file = "result.xlsx"

participant_name, result = evaluate_answers(file_path, sheet_name, row_number)
save_results(output_file, participant_name, result)
