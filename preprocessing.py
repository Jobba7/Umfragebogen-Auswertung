from openpyxl import load_workbook, Workbook


def extract_participant_responses(sheet, row_number):
    # Teilnehmername aus Spalte H
    participant_name = sheet[f"H{row_number}"].value

    # Startspalte für die Antworten (Frage 2 beginnt bei Spalte T)
    start_column = ord("T")  # Spalte T als Basis
    total_questions = 28
    answer_offset = 7  # Abstand der Fragen
    number_of_answer_options = 3  # Anzahl der Antwortmöglichkeiten (A, B, C)

    answers = []  # Hier speichern wir die Antworten ('A', 'B' oder 'C')

    # Schleife über alle Fragen
    for question in range(1, total_questions + 1):
        # Spalten der Antwortmöglichkeiten (A, B, C)
        answer_columns = []
        for i in range(number_of_answer_options):
            answer_column_index = start_column + (question - 1) * answer_offset + i
            first_letter = chr(ord("A") + (answer_column_index - ord("A")) // 26 - 1)
            second_letter = chr(ord("A") + (answer_column_index - ord("A")) % 26)
            if first_letter < "A":
                answer_columns.append(second_letter)
            else:
                answer_columns.append(first_letter + second_letter)

        # Werte der Zellen abrufen
        values = [sheet[f"{col}{row_number}"].value for col in answer_columns]

        # Prüfen, welche Antwort gewählt wurde (1 steht für die Auswahl)
        if values[0] == 1:
            answers.append("A")
        elif values[1] == 1:
            answers.append("B")
        elif values[2] == 1:
            answers.append("C")
        else:
            answers.append(None)  # Für den Fall, dass keine gültige Antwort vorliegt

    # Rückgabe des Teilnehmernamens und der Antworten
    return participant_name, answers


def process_excel_data(input_file):
    # Excel-Datei mit den Daten laden
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Neue Excel-Datei erstellen
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Results"

    # Header hinzufügen
    output_sheet.append(["Teilnehmer", *[f"Frage {i+1}" for i in range(28)]])

    # Verarbeitung aller Teilnehmer
    start_row = 3
    current_row = start_row
    while True:
        # Teilnehmername aus der aktuellen Zeile
        participant_name = sheet[f"H{current_row}"].value

        # Stop, wenn eine leere Zeile erreicht wird
        if not participant_name:
            break

        # Antworten auswerten
        _, answers = extract_participant_responses(sheet, current_row)

        # Ergebnisse speichern
        output_sheet.append([participant_name, *answers])

        # Zur nächsten Zeile wechseln
        current_row += 1

    workbook.close()
    return output_workbook
